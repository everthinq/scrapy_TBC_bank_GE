import json
import scrapy
from openpyxl import load_workbook


class TBC_bank(scrapy.Spider):
    name = 'TBC_bank'

    def start_requests(self):
        self.cell_value = '2'
        self.workbook = load_workbook('JSC TBC Bank Georgia.xlsx')
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

        for id in range(0, 5000):
            post_body = 'poiId=' + str(id)
            yield scrapy.Request(
                method='POST',
                url='https://www.tbcbank.ge/web/en/web/guest/branches-and-atms?p_p_id=atm_WAR_tbcpwatmportlet&p_p_lifecycle=2&p_p_state=normal&p_p_mode=view&p_p_resource_id=getPoiDetail&p_p_cacheability=cacheLevelPage&p_p_col_id=column-1&p_p_col_pos=2&p_p_col_count=3',
                headers={'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'},
                body=post_body,
                callback=self.parse_json)
            # break

    def parse_json(self, response):
        if (response.text):
            JSON = json.loads(response.text)
            # print(JSON)

            if JSON is not None and JSON['typeNameEng'] == 'Branch':
                branch_name = JSON['nameEng'].strip()
                address = JSON['addressEng']
                region = JSON['regionNameEng']
                lat = JSON['position']['latitude']
                lng = JSON['position']['longitude']

                print('Writing --', 'name:', branch_name, 'address:', address, 'region:', region,
                      'lat:', lat, 'lng:', lng)

                self.cell_value = str(self.cell_value)
                self.worksheet['B' + self.cell_value] = 'JSC TBC Bank Georgia'
                self.worksheet['C' + self.cell_value] = branch_name
                self.worksheet['D' + self.cell_value] = address
                self.worksheet['G' + self.cell_value] = 'Georgia'
                self.worksheet['H' + self.cell_value] = 'GE'
                self.worksheet['I' + self.cell_value] = region
                self.worksheet['M' + self.cell_value] = lat
                self.worksheet['N' + self.cell_value] = lng
                self.worksheet['O' + self.cell_value] = 'Address'
                self.worksheet['R' + self.cell_value] = 'Bank website'
                self.cell_value = int(self.cell_value) + 1

        self.workbook.save('JSC TBC Bank Georgia.xlsx')
